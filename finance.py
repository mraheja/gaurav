import requests
import pandas
import openpyxl

def read_tickers(filename): #filename in xlsx presumably
    sheet = openpyxl.load_workbook(filename).get_sheet_by_name("Sheet1")
    return [x.value for x in sheet["A"]][1:]

print(read_tickers("listcode.xlsx"))

def add_new_prices(prices, filename, new_file):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name("Sheet1")
    columns = sheet.max_column + 1
    for row, price in enumerate(prices, 1): 
        sheet.cell(row=row, column=columns, value=price)
    workbook.save(new_file)            
url = 'https://www.alphavantage.co/query'
tickers = ['MSFT','GOOG','X']

for sym in tickers:
	params = dict(
		function='GLOBAL_QUOTE',
		symbol=sym,
		apikey='ZCQPZJ9MFTUI4OX6'
	)

	resp = requests.get(url=url, params=params)
	data = resp.json()
	print(data["Global Quote"]["05. price"])
